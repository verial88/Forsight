from pyexcel_xlsx import get_data
from openpyxl import load_workbook

# метод с использованием библиотеки pyexcel_xlsx
# data = get_data('example_tree1.xlsx')
# print(data)
# import json
# print(json.dumps(data))

# метод с использованием библиотеки openpyxl
wb_obj = load_workbook(filename='example_tree1.xlsx', data_only=True)  #data_only режим вывода только информации (значения, а не формулы)
sheet_obj = wb_obj['Лист1']
# B1_obj = sheet_obj['B1'].value
# print(B1_obj)
# sheet_obj = wb_obj.active
# print(sheet_obj)
# cell_obj = sheet_obj.cell(row=1, column=2)
# print(cell_obj.value)
# print(sheet_obj.max_column)
# print(sheet_obj.max_row)

# list_v = []
# for i in range(1, sheet_obj.max_row + 1):
#     for j in range(1, sheet_obj.max_column + 1):
#         list_v.append(sheet_obj.cell(row=i, column=j).value)
#     print(list_v)

# for row in sheet_obj.values:
#     for value in row:
#         print(value)

for row in sheet_obj.iter_rows(min_row=1, max_col=3, values_only=True):
 if 'класс22' in row:
  print(row[0])


